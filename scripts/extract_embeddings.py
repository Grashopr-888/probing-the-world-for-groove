#!/usr/bin/env python3
"""
Extract within-model primary-style centroid-shift data for the Representations page.

Source : public/data/thesis-experiment-results.xlsx
  sheets: GMD_CNN_primary_embedding_compa  (CNN, "Mean Primary Style Shift")
          PaSST_primary_embedding_compari  (PaSST, "Mean Primary Style Centroid Shift")
Output : src/data/embedding-comparisons.json

IMPORTANT: CNN and PaSST shifts live in different coordinate systems with different
scales — they are emitted as SEPARATE within-model series and must never be plotted
on a shared axis. Each sheet is pre-sorted by descending mean shift; row 1 is the
largest representational change observed for that model.
"""
import json
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "public" / "data" / "thesis-experiment-results.xlsx"
OUT = ROOT / "src" / "data" / "embedding-comparisons.json"

SHEETS = {"cnn": "GMD_CNN_primary_embedding_compa", "passt": "PaSST_primary_embedding_compari"}
TOP_N = 10


def extract(ws):
    rows = [r for r in ws.iter_rows(values_only=True) if r and r[0] is not None]
    rows = rows[1:]  # drop header
    # first distinct (A,B) pair == largest mean shift (sheet is sorted desc)
    a, b, mean = rows[0][0], rows[0][1], round(float(rows[0][2]), 2)
    pair = [r for r in rows if r[0] == a and r[1] == b]
    desc = pair[0][6] if len(pair[0]) > 6 and pair[0][6] else ""
    f1a = round(float(pair[0][7]), 4) if len(pair[0]) > 7 and pair[0][7] is not None else None
    f1b = round(float(pair[0][8]), 4) if len(pair[0]) > 8 and pair[0][8] is not None else None
    top = [{"style": str(r[4]), "shift": round(float(r[5]), 2)} for r in pair[:TOP_N]]
    return {
        "expA": a, "expB": b, "description": desc, "f1A": f1a, "f1B": f1b,
        "meanPrimaryShift": mean, "topPrimary": top,
    }


CNN_LABELS = {
    "GMD_CNN_prototype6_7": "Reflection padding", "GMD_CNN_prototype6_8": "Circular padding",
    "GMD_CNN_prototype6_5": "Time stretch", "GMD_CNN_prototype6_4": "Noise + room",
    "GMD_CNN_prototype6_6": "Noise, room + reflection",
}
PASST_LABELS = {
    "PaSST_setup6_15": "Circular padding", "PaSST_setup6_14": "Time stretch",
    "PaSST_setup6_12": "Deeper head",
}


def interventions(ws, baseline, labels):
    """Mean primary-style centroid shift from a baseline run to each labelled intervention."""
    rows = [r for r in ws.iter_rows(values_only=True) if r and r[0] is not None][1:]
    seen = {}
    for r in rows:
        a, b = str(r[0]), str(r[1])
        if a == baseline and b in labels and b not in seen:
            seen[b] = round(float(r[2]), 2)
    return sorted(({"label": labels[b], "shift": s} for b, s in seen.items()),
                  key=lambda d: -d["shift"])


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    data = {m: {"model": m, **extract(wb[s])} for m, s in SHEETS.items()}
    data["cnn"]["interventionBaseline"] = "zero padding"
    data["cnn"]["interventionShifts"] = interventions(wb[SHEETS["cnn"]], "GMD_CNN_prototype6_2", CNN_LABELS)
    data["passt"]["interventionBaseline"] = "reflection padding"
    data["passt"]["interventionShifts"] = interventions(wb[SHEETS["passt"]], "PaSST_setup6_16", PASST_LABELS)
    data["note"] = (
        "CNN and PaSST embeddings occupy different coordinate systems with different distance "
        "scales. The two series below are within-model only and share no axis."
    )
    data["scaleRatio"] = round(data["cnn"]["meanPrimaryShift"] / data["passt"]["meanPrimaryShift"], 1)
    OUT.write_text(json.dumps(data, indent=2))
    print(f"✓ embedding-comparisons.json")
    print(f"  CNN mean primary shift {data['cnn']['meanPrimaryShift']} ({data['cnn']['expA']}→{data['cnn']['expB']})")
    print(f"  PaSST mean primary shift {data['passt']['meanPrimaryShift']} ({data['passt']['expA']}→{data['passt']['expB']})")
    print(f"  scale ratio ≈ {data['scaleRatio']}×")
    print(f"  CNN top: {', '.join(t['style'] for t in data['cnn']['topPrimary'][:3])} …")
    print(f"  PaSST top: {', '.join(t['style'] for t in data['passt']['topPrimary'][:3])} …")


if __name__ == "__main__":
    main()
